# d:\AAA_Framework_Acceso\ProjectAcceso\appacceso\admin.py
from django.contrib import admin
from .models import RegistroFirma, Sede
from django.utils.html import format_html
from import_export import resources
from semantic_admin.contrib.import_export.admin import SemanticImportExportModelAdmin
from django.http import HttpResponse
from django.utils import timezone # Para el nombre del archivo
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
import openpyxl.styles # Para estilos de celda como negrita
from io import BytesIO
from PIL import Image as PillowImage # Para abrir y manipular la imagen, Pillow es una dependencia de ImageField
import os # Para verificar si el archivo de imagen existe


@admin.register(Sede)
class SedeAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'direccion', 'ciudad', 'telefono')
    search_fields = ('nombre', 'ciudad')
    list_filter = ('ciudad',)

class RegistroFirmaResource(resources.ModelResource):
    # Personalizar los campos y cómo se exportan
    usuario_username = resources.Field(attribute='usuario__username', column_name='Usuario')
    sede_nombre = resources.Field(attribute='sede__nombre', column_name='Sede')
    fecha_ingreso_formateada = resources.Field(attribute='fecha_ingreso', column_name='Fecha de Ingreso')
    fecha_grabacion_formateada = resources.Field(attribute='fecha_grabacion', column_name='Fecha de Grabación')
    # Si quieres exportar la URL de la firma (ten en cuenta que esto es solo la URL, no la imagen en sí)
    firma_url = resources.Field(column_name='URL Firma') # Activamos este campo


    class Meta:
        model = RegistroFirma
        # Define los campos que quieres exportar y su orden
        fields = ('id', 'usuario_username', 'sede_nombre', 'fecha_ingreso_formateada', 'fecha_grabacion_formateada', 'firma_url') # Añadimos firma_url
        export_order = fields # Mantiene el orden definido en fields

    def dehydrate_fecha_ingreso_formateada(self, registro):
        return registro.fecha_ingreso.strftime("%Y-%m-%d %H:%M:%S") if registro.fecha_ingreso else ''
    
    def dehydrate_fecha_grabacion_formateada(self, registro):
        return registro.fecha_grabacion.strftime("%Y-%m-%d %H:%M:%S") if registro.fecha_grabacion else ''
    
    # Si decides exportar la URL de la firma:
    def dehydrate_firma_url(self, registro): # Activamos este método
        if registro.firma:
            return registro.firma.url # Esto da la URL relativa, ej: /media/firmas/2023/05/09/firma.png
        return ''

@admin.register(RegistroFirma)
class RegistroFirmaAdmin(SemanticImportExportModelAdmin): # Cambiar la herencia a la de django-import-export
    # Forzar el uso de nuestra plantilla sobrescrita
    resource_classes = [RegistroFirmaResource] # Usar la clase Resource
    list_display = ('usuario', 'sede', 'fecha_ingreso', 'fecha_grabacion', 'ver_firma')
    list_filter = ('fecha_ingreso', 'usuario', 'sede')
    search_fields = ('usuario__username', 'sede__nombre')
    # Hacemos 'sede' editable en el form de admin si no es readonly
    readonly_fields = ('usuario', 'fecha_ingreso', 'fecha_grabacion', 'firma_preview')
    # Si quieres que 'sede' sea editable en el admin, quítalo de readonly_fields.
    # Si 'sede' es null=True en el modelo y quieres que sea opcional en el admin, está bien.
    # Si 'sede' es null=False, debe ser un campo requerido.

    fieldsets = (
        (None, {
            'fields': ('usuario', 'sede', 'fecha_ingreso', 'firma_preview', 'fecha_grabacion')
        }),
    )
    # La acción de exportar se añade automáticamente por ImportExportModelAdmin

    def ver_firma(self, obj):
        if obj.firma and hasattr(obj.firma, 'url'):
            return format_html('<a href="{}" target="_blank"><img src="{}" width="100" height="50" style="object-fit: contain;" /></a>', obj.firma.url, obj.firma.url)
        return "Sin firma"
    ver_firma.short_description = 'Firma (Click para ampliar)'

    def firma_preview(self, obj): # Para el form de admin
        if obj.firma:
            return format_html('<img src="{}" width="300" height="150" style="object-fit: contain;" />', obj.firma.url)
        return "No hay firma adjunta."
    firma_preview.short_description = 'Vista Previa Firma'

    # Nueva acción para exportar con imágenes incrustadas
    def export_selected_to_excel_with_images(self, request, queryset):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="registros_firmas_con_imagenes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Registros de Firmas'

        # Definir cabeceras
        headers = ['ID', 'Usuario', 'Sede', 'Fecha de Ingreso', 'Fecha de Grabación', 'Firma']
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header_title)
            cell.font = openpyxl.styles.Font(bold=True)

        # Escribir datos
        row_num = 2
        # Altura de la imagen en píxeles y altura de la fila en puntos
        # 1 píxel ~= 0.75 puntos (para 96 DPI pantalla / 72 DPI Excel)
        img_display_height_px = 80  # Altura deseada para la imagen en la celda (en píxeles)
        row_height_pt = (img_display_height_px * 0.75) + 15 # Altura de la fila en puntos, con padding

        # Ancho máximo de la imagen en píxeles, para evitar que sea demasiado ancha
        img_max_width_px = 200 
        # Ancho de la columna de la firma (aproximado en unidades de caracteres de Excel)
        # 1 unidad de ancho de carácter ~= 7 píxeles (con fuente Calibri 11)
        firma_col_width = (img_max_width_px / 7) + 2 

        for registro in queryset:
            sheet.cell(row=row_num, column=1, value=registro.id)
            sheet.cell(row=row_num, column=2, value=registro.usuario.username)
            sheet.cell(row=row_num, column=3, value=registro.sede.nombre if registro.sede else 'N/A')
            sheet.cell(row=row_num, column=4, value=registro.fecha_ingreso.strftime("%Y-%m-%d %H:%M:%S") if registro.fecha_ingreso else '')
            sheet.cell(row=row_num, column=5, value=registro.fecha_grabacion.strftime("%Y-%m-%d %H:%M:%S") if registro.fecha_grabacion else '')

            firma_cell_anchor = f'{get_column_letter(6)}{row_num}'
            sheet.row_dimensions[row_num].height = row_height_pt # Establecer altura de fila

            if registro.firma and hasattr(registro.firma, 'path') and registro.firma.path and os.path.exists(registro.firma.path):
                try:
                    img_pil = PillowImage.open(registro.firma.path)
                    original_width, original_height = img_pil.size
                    aspect_ratio = original_width / original_height

                    # Calcular ancho de la imagen basado en la altura fija y la relación de aspecto
                    current_img_width_px = int(img_display_height_px * aspect_ratio)
                    current_img_height_px = img_display_height_px

                    # Si la imagen calculada es más ancha que el máximo, reescalar por ancho
                    if current_img_width_px > img_max_width_px:
                        current_img_width_px = img_max_width_px
                        current_img_height_px = int(current_img_width_px / aspect_ratio)

                    img_openpyxl = OpenpyxlImage(registro.firma.path)
                    img_openpyxl.height = current_img_height_px
                    img_openpyxl.width = current_img_width_px
                    
                    sheet.add_image(img_openpyxl, firma_cell_anchor)
                except FileNotFoundError:
                    sheet.cell(row=row_num, column=6, value="Archivo no encontrado")
                except Exception: # Captura otros errores de Pillow/Openpyxl
                    sheet.cell(row=row_num, column=6, value="Error al cargar imagen")
            else:
                sheet.cell(row=row_num, column=6, value="Sin firma")
            
            row_num += 1

        # Ajustar anchos de columna (excepto la de la imagen que ya tiene un ancho estimado)
        for i, column_cells in enumerate(sheet.columns):
            col_letter = get_column_letter(i + 1)
            if col_letter == get_column_letter(6): # Columna de la firma
                sheet.column_dimensions[col_letter].width = firma_col_width
                continue
            
            max_length = 0
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_length + 2
        
        workbook.save(response)
        return response

    export_selected_to_excel_with_images.short_description = "Exportar seleccionados a Excel (con imágenes)"

    actions = [export_selected_to_excel_with_images] # Añadir la acción al desplegable
